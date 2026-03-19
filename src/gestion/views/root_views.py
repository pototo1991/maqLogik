from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from ..models import Empresa
import logging

logger = logging.getLogger('gestion')

def is_superadmin(user):
    """ Verifica que el usuario tenga privilegios de Root SaaS """
    return user.is_authenticated and user.is_superuser

@user_passes_test(is_superadmin, login_url='/')
def root_dashboard(request):
    """ Panel de Control Maestro para Administración de MaqLogik SaaS """
    empresas = Empresa.objects.all().order_by('-id')
    
    # KPIs Rápidos
    total_empresas = empresas.count()
    
    # Data para Gráficos
    empresas_activas = empresas.filter(activa=True).count()
    empresas_inactivas = empresas.filter(activa=False).count()
    
    modulos_data = [
        empresas.filter(modulo_gps=True).count(),
        empresas.filter(modulo_mantencion=True).count(),
        empresas.filter(modulo_combustible=True).count(),
        empresas.filter(modulo_checklist=True).count(),
        empresas.filter(modulo_reporteria=True).count()
    ]
    
    context = {
        'empresas': empresas,
        'total_empresas': total_empresas,
        'empresas_activas': empresas_activas,
        'empresas_inactivas': empresas_inactivas,
        'modulos_data': modulos_data,
    }
    return render(request, 'gestion/root/dashboard.html', context)

from django.db import transaction
from ..forms import EmpresaForm, UsuarioCreationForm, ConfiguracionMantencionForm
from ..models import Usuario, ConfiguracionMantencion

@user_passes_test(is_superadmin, login_url='/')
def root_create_empresa(request):
    """ Vista para crear una Nueva Empresa con sus respectivos módulos """
    if request.method == 'POST':
        form_empresa = EmpresaForm(request.POST)
        form_owner = UsuarioCreationForm(request.POST)
        form_mantencion = ConfiguracionMantencionForm(request.POST)
        
        if form_empresa.is_valid() and form_owner.is_valid() and form_mantencion.is_valid():
            try:
                with transaction.atomic():
                    # 1. Crear Empresa
                    nueva_empresa = form_empresa.save()
                    
                    # 1.5 Crear Configuracion Mantencion
                    nueva_config = form_mantencion.save(commit=False)
                    nueva_config.empresa = nueva_empresa
                    nueva_config.save()
                    
                    # 2. Asignarle el formulario de Owner a esa empresa
                    nuevo_owner = form_owner.save(commit=False)
                    nuevo_owner.empresa = nueva_empresa
                    nuevo_owner.rol = 'OWNER'
                    nuevo_owner.save()
                    
                logger.info(f"AUDITORÍA - Nueva Empresa SaaS Creada: '{nueva_empresa.nombre_fantasia}' (RUT: {nueva_empresa.rut_empresa}) por el administrador {request.user.username}.")
                messages.success(request, f"¡Empresa {nueva_empresa.nombre_fantasia} creada con éxito y Owner asignado!")
                return redirect('root_dashboard')
            except Exception as e:
                logger.error(f"AUDITORÍA - EXCEPCIÓN DB AL CREAR EMPRESA: {str(e)}")
                messages.error(request, f"Error interno al crear Empresa/Owner: {str(e)}")
        else:
            errors_msg = "Revise: "
            if form_empresa.errors:
                errors_msg += f"Empresa: {form_empresa.errors.as_text()} | "
            if form_owner.errors:
                errors_msg += f"Owner: {form_owner.errors.as_text()} | "
            if form_mantencion.errors:
                errors_msg += f"Mantención: {form_mantencion.errors.as_text()}"
            
            # Formateando y enviando los errores exactos a los logs del archivo app_control.log
            logger.warning(f"AUDITORÍA - Falla de validación de Formulario al crear empresa: {errors_msg.replace('  ', ' ')}")
            messages.error(request, errors_msg)
    else:
        form_empresa = EmpresaForm()
        form_owner = UsuarioCreationForm()
        form_mantencion = ConfiguracionMantencionForm()

        
    context = {
        'form_empresa': form_empresa,
        'form_owner': form_owner,
        'form_mantencion': form_mantencion
    }
    return render(request, 'gestion/root/nueva_empresa.html', context)

@user_passes_test(is_superadmin, login_url='/')
def root_edit_modules(request, empresa_id):
    """ Vista para editar los módulos contratados de una Empresa (SaaS Features) """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    config_mantencion, created = ConfiguracionMantencion.objects.get_or_create(empresa=empresa)
    
    if request.method == 'POST':
        # Le pasamos la instancia existente para que haga Update en lugar de Create
        form = EmpresaForm(request.POST, instance=empresa)
        form_mantencion = ConfiguracionMantencionForm(request.POST, instance=config_mantencion)
        if form.is_valid() and form_mantencion.is_valid():
            form.save()
            form_mantencion.save()
            messages.success(request, f"¡Módulos de la empresa {empresa.nombre_fantasia} actualizados correctamente!")
            return redirect('root_dashboard')
        else:
            messages.error(request, "Error al actualizar los módulos. Verifique los campos.")
    else:
        form = EmpresaForm(instance=empresa)
        form_mantencion = ConfiguracionMantencionForm(instance=config_mantencion)
        
    context = {
        'form': form,
        'form_mantencion': form_mantencion,
        'empresa': empresa
    }
    return render(request, 'gestion/root/editar_modulos.html', context)

from django.contrib.auth.forms import SetPasswordForm

@user_passes_test(is_superadmin, login_url='/')
def root_manage_owner(request, empresa_id):
    """ Vista para gestionar credenciales y password del Owner local de la Empresa """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    # IMPORTANTE: Usamos la relación inversa (empresa.usuario_set) en lugar de Usuario.objects
    # porque el UsuarioManager custom filtra por empresa en sesión y el Superadmin no tiene empresa.
    owner = empresa.usuario_set.filter(rol='OWNER').first()
    
    if not owner:
        messages.error(request, f"La empresa {empresa.nombre_fantasia} aún no tiene un Owner asignado.")
        return redirect('root_dashboard')
        
    if request.method == 'POST':
        # Instanciamos SetPasswordForm (requerido para resetar password ignorando la antigua por parte un admin)
        form = SetPasswordForm(user=owner, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"Contraseña del Owner '{owner.username}' actualizada exitosamente.")
            return redirect('root_dashboard')
        else:
            messages.error(request, "Vuelva a verificar que las contraseñas coinciden.")
    else:
        # Pasa el usuario Owner
        form = SetPasswordForm(user=owner)
        
    context = {
        'form': form,
        'empresa': empresa,
        'owner': owner
    }
    return render(request, 'gestion/root/gestionar_owner.html', context)

@user_passes_test(is_superadmin, login_url='/')
def root_impersonate(request, empresa_id):
    """ 
    Permite al Superusuario 'Ver como' si fuera un empleado de esta Empresa.
    Guarda el ID temporalmente en la sesión, el cual debe ser leído por el Middleware.
    """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    request.session['impersonate_empresa_id'] = empresa.id
    messages.info(request, f"👀 Estás viendo el sistema como: {empresa.nombre_fantasia}")
    return redirect('dashboard')  # Enviamos al dashboard principal de la app

@user_passes_test(is_superadmin, login_url='/')
def root_stop_impersonate(request):
    """ Sale del modo 'Ver como' y destruye la variable de sesión """
    if 'impersonate_empresa_id' in request.session:
        del request.session['impersonate_empresa_id']
        messages.success(request, "Has regresado al Panel de Control Maestro (SaaS).")
    return redirect('root_dashboard')

@user_passes_test(is_superadmin, login_url='/')
def root_toggle_empresa_status(request, empresa_id):
    """ 
    Invierte el estado de 'activa' de una empresa (Soft Delete/Restore).
    También desactiva masivamente todos los usuarios asociados para denegar logins,
    excluyendo expresamente a los superusuarios.
    """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    usuarios_empresa = Usuario.objects.filter(empresa=empresa)
    
    # Toggle del booleano
    nuevo_estado = not empresa.activa
    empresa.activa = nuevo_estado
    empresa.save()

    # Actualizar is_active forzosamente a todos los empleados conectados a la empresa
    # EXCEPTO al panel maestro / superusuarios para evitar bloqueos del sistema.
    usuarios_empresa.exclude(is_superuser=True).update(is_active=nuevo_estado)

    # Mensajes UI
    accion = "Habilitada" if nuevo_estado else "Deshabilitada (Soft Delete aplicado)"
    color = messages.success if nuevo_estado else messages.warning
    color(request, f"La empresa {empresa.nombre_fantasia} fue {accion}.")
    
    return redirect('root_dashboard')

import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse

@user_passes_test(is_superadmin, login_url='/')
def root_download_plantilla_usuarios(request, empresa_id):
    """ Genera y descarga una plantilla Excel (.xlsx) con los encabezados correctos """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    headers = ['RUT', 'Nombres_y_Apellidos', 'Correo', 'Telefono', 'Rol', 'Valor_Hora']
    ws.append(headers)
    
    # Estilos cabecera
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        
    ws.append(['11111111-1', 'Juan Perez', 'juan@ejemplo.com', '912345678', 'OPERATOR', 0])
    ws.append(['22222222-2', 'Pedro (Jefe Patio)', 'pedro@ejemplo.com', '987654321', 'CHIEF', 15000])
    
    # Ajuste ancho columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_usuarios.xlsx"'
    wb.save(response)
    
    return response

@user_passes_test(is_superadmin, login_url='/')
def root_import_usuarios(request, empresa_id):
    """ Módulo para procesar subida masiva Excel (.xlsx) de Usuarios desde el Admin SaaS """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un Excel .xlsx')
            return redirect('root_import_usuarios', empresa_id=empresa.id)
            
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            usuarios_creados = 0
            errores = []
            
            roles_validos = dict(Usuario.ROLES).keys()

            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: # Fila en blanco o sin RUT
                        continue
                        
                    rut = str(row[0]).strip()
                    nombre_completo = str(row[1]).strip() if row[1] else ""
                    correo = str(row[2]).strip() if row[2] else ""
                    telefono = str(row[3]).strip() if row[3] else ""
                    
                    # Normalización automática de teléfono (clásica petición de Chile)
                    if telefono:
                        # Si solo pusieron los 9 dígitos, agregamos el +56
                        if len(telefono) == 9 and telefono.startswith('9'):
                            telefono = f"+56{telefono}"
                        # Eliminar posibles espacios o caracteres extraños que Excel a veces mete
                        telefono = telefono.replace(" ", "")

                    rol = str(row[4]).strip().upper() if row[4] else ""
                    
                    try:
                        valor_hora = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0
                    except (ValueError, TypeError):
                        valor_hora = 0.0
                        
                    if rol not in roles_validos:
                        errores.append(f"Fila {row_idx}: Rol '{rol}' no es válido. Opciones: {', '.join(roles_validos)}")
                        continue
                        
                    if Usuario.objects.filter(rut=rut).exists():
                        errores.append(f"Fila {row_idx}: El RUT {rut} ya existe.")
                        continue
                        
                    # Generar username desde nombre (idéntico al formulario manual y JS)
                    import unicodedata
                    nombre_norm = unicodedata.normalize('NFD', nombre_completo.lower())
                    nombre_norm = ''.join(c for c in nombre_norm if unicodedata.category(c) != 'Mn')
                    username_base = nombre_norm.strip().replace(' ', '_')
                    username_base = ''.join(c for c in username_base if c.isalnum() or c == '_')
                    if not username_base:  # Fallback si el nombre estaba vacío
                        username_base = rut.replace('-', '').replace('.', '')
                    if Usuario.objects.filter(username=username_base).exists():
                        rut_sufijo = rut.replace('.', '').replace('-', '')[-4:]
                        username_base = f"{username_base}_{rut_sufijo}"
                        
                    nuevo_usuario = Usuario(
                        username=username_base,
                        rut=rut,
                        nombre_completo=nombre_completo,
                        email=correo,
                        telefono=telefono,
                        rol=rol,
                        valor_hora=valor_hora,
                        empresa=empresa,
                        is_active=True,
                        debe_cambiar_password=True  # Fuerza cambio en 1er login
                    )
                    clave_temporal = rut.replace('.', '').replace('-', '')
                    nuevo_usuario.set_password(clave_temporal)
                    nuevo_usuario.save()
                    usuarios_creados += 1
                    
            if errores:
                for error in errores[:5]:
                    messages.warning(request, error)
                if len(errores) > 5:
                    messages.warning(request, f"... y {len(errores)-5} errores adicionales en el archivo.")
            
            if usuarios_creados > 0:
                messages.success(request, f"Se importaron {usuarios_creados} empleados a {empresa.nombre_fantasia}.")
                logger.info(f"AUDITORÍA - Importación Masiva Excel: {usuarios_creados} usuarios a '{empresa.nombre_fantasia}'.")
                return redirect('root_dashboard')
                
        except Exception as e:
            logger.error(f"Error procesando Excel Import Usuarios: {str(e)}")
            messages.error(request, f"Error leyendo Excel. Detalle: {str(e)}")
            
    return render(request, 'gestion/root/importar_usuarios.html', {'empresa': empresa})

from django.core.exceptions import ObjectDoesNotExist
from ..models import Maquinaria

@user_passes_test(is_superadmin, login_url='/')
def root_download_plantilla_maquinarias(request, empresa_id):
    """ Genera y descarga plantilla Excel (.xlsx) para Maquinarias """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flota"
    
    headers = ['ID_Interno', 'Patente', 'Tipo', 'Marca', 'Modelo', 'Tonelaje', 'Combustible', 'Unidad_Medida', 'Valor_Actual', 'Consumo_Teorico', 'Proximo_Mantenimiento']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        
    ws.append(['EXC-01', 'ABCD12', 'EXCAVADORA', 'CATERPILLAR', '320', 20.0, 'DIESEL', 'HORAS', 5000, 15.5, 5250])
    ws.append(['CAM-01', 'XYZW99', 'CAMION', 'VOLVO', 'FH', 0.0, 'DIESEL', 'KM', 120000, 3.2, 130000])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_maquinarias.xlsx"'
    wb.save(response)
    return response

@user_passes_test(is_superadmin, login_url='/')
def root_import_maquinarias(request, empresa_id):
    """ Módulo para procesar subida masiva Excel (.xlsx) de Maquinarias """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un Excel .xlsx')
            return redirect('root_import_maquinarias', empresa_id=empresa.id)
            
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            maquinas_creadas = 0
            errores = []
            
            # Obtener configuración de mantención de la empresa para autocálculos
            try:
                config_mant = empresa.configuracion_mantencion
                default_horas = config_mant.intervalo_horas
                default_km = config_mant.intervalo_km
            except ObjectDoesNotExist:
                default_horas = 250
                default_km = 10000
            
            tipos_validos = dict(Maquinaria.TIPOS_MAQUINA).keys()
            unidades_validas = dict(Maquinaria.UNIDADES).keys()  # HORAS, KM
            # Combustible es campo libre en el modelo; definimos los valores aceptados aquí
            combs_validos = {'DIESEL', 'BENCINA', 'GLP', 'ELECTRICO', 'OTRO'}

            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[1]: # Fila sin Patente
                        continue
                        
                    id_interno = str(row[0]).strip() if row[0] else ""
                    patente = str(row[1]).strip().upper()
                    tipo = str(row[2]).strip().upper() if row[2] else ""
                    marca = str(row[3]).strip().upper() if row[3] else ""
                    modelo = str(row[4]).strip().upper() if row[4] else ""
                    
                    try: tonelaje = float(row[5]) if row[5] is not None else 0.0
                    except: tonelaje = 0.0
                    
                    combustible = str(row[6]).strip().upper() if len(row) > 6 and row[6] else "DIESEL"
                    medida = str(row[7]).strip().upper() if len(row) > 7 and row[7] else "HORAS"
                    
                    try: valor_actual = float(row[8]) if len(row) > 8 and row[8] is not None else 0.0
                    except: valor_actual = 0.0
                    
                    try: consumo = float(row[9]) if len(row) > 9 and row[9] is not None else 0.0
                    except: consumo = 0.0
                    
                    # AUTOCÁLCULO OP. 2: Híbrido
                    try:
                        prox_mant = float(row[10]) if len(row) > 10 and row[10] is not None else None
                    except:
                        prox_mant = None
                        
                    # Si viene vacío o nulo (None o ''), calculamos
                    if prox_mant is None:
                        if medida == 'KM':
                            prox_mant = valor_actual + default_km
                        else:
                            prox_mant = valor_actual + default_horas

                    if tipo not in tipos_validos:
                        errores.append(f"Fila {row_idx}: Tipo '{tipo}' no válido.")
                        continue
                    if combustible not in combs_validos:
                        errores.append(f"Fila {row_idx}: Combustible '{combustible}' no válido.")
                        continue
                    if medida not in unidades_validas:
                        errores.append(f"Fila {row_idx}: Medida '{medida}' no válida.")
                        continue
                        
                    if Maquinaria.objects.filter(patente=patente).exists():
                        errores.append(f"Fila {row_idx}: Patente {patente} ya existe.")
                        continue
                         
                    nueva_maq = Maquinaria(
                        empresa=empresa,
                        id_interno=id_interno,
                        patente=patente,
                        tipo=tipo,
                        marca=marca,
                        modelo=modelo,
                        tonelaje=tonelaje,
                        tipo_combustible=combustible,
                        unidad_medida=medida,
                        valor_actual_medida=valor_actual,
                        consumo_teorico=consumo,
                        proximo_mantenimiento=prox_mant,
                        estado='DISPONIBLE',
                        operador_asignado=None # Solicitado por el usuario: En blanco por defecto
                    )
                    nueva_maq.save()
                    maquinas_creadas += 1
                    
            if errores:
                for error in errores[:5]:
                    messages.warning(request, error)
                if len(errores) > 5:
                    messages.warning(request, f"... y {len(errores)-5} errores adicionales en el archivo.")
            
            if maquinas_creadas > 0:
                messages.success(request, f"Se importaron {maquinas_creadas} maquinarias a {empresa.nombre_fantasia}.")
                logger.info(f"AUDITORÍA - Importación Excel: {maquinas_creadas} máquinas a '{empresa.nombre_fantasia}'.")
                return redirect('root_dashboard')
                
        except Exception as e:
            logger.error(f"Error procesando Excel Import Maquinarias: {str(e)}")
            messages.error(request, f"Error leyendo Excel. Detalle: {str(e)}")
            
    return render(request, 'gestion/root/importar_maquinarias.html', {'empresa': empresa})
